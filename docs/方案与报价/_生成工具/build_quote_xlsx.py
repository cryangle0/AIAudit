#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
爱对 - 多平台店铺财务自动对账系统 报价单 (xlsx) 生成器

预算：5 万元
周期：20 个工作日

参考自 E:\\angsa\\angsa_data\\AITools\\create_xlsx.py
"""

import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


# ==================== 配色 ====================
COLORS = {
    'main_title_bg': '2C5F9E',    # 深蓝
    'main_title_font': 'FFFFFF',
    'header_bg': 'D9E8F5',
    'header_font': '333333',
    'module_title_bg': 'F2F2F2',
    'module_title_font': '2C5F9E',
    'subtotal_bg': 'FFF9E6',
    'subtotal_font': '333333',
    'total_bg': '2C5F9E',
    'total_font': 'FFFFFF',
    'section_title_bg': 'E8F4FD',
    'section_title_font': '2C5F9E',
}

THIN_BORDER = Border(
    left=Side(style='thin', color='999999'),
    right=Side(style='thin', color='999999'),
    top=Side(style='thin', color='999999'),
    bottom=Side(style='thin', color='999999')
)
THICK_BORDER = Border(
    left=Side(style='medium', color='2C5F9E'),
    right=Side(style='medium', color='2C5F9E'),
    top=Side(style='medium', color='2C5F9E'),
    bottom=Side(style='medium', color='2C5F9E')
)


class ProjectQuoteGenerator:
    """通用项目报价单生成器（保留与参考脚本一致的接口）"""

    def __init__(self, project_name, sheet_name=None):
        self.project_name = project_name
        self.sheet_name = sheet_name or f'{project_name}报价单'
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = self.sheet_name[:31]
        self.current_row = 1
        self.subtotal_rows = []
        self._setup_columns()

    def _setup_columns(self):
        widths = {'A': 22, 'B': 20, 'C': 50, 'D': 60, 'E': 10, 'F': 12, 'G': 10}
        for col, w in widths.items():
            self.ws.column_dimensions[col].width = w

    def _merge_and_style(self, row, start_col, end_col, value, font, fill=None,
                        alignment=None, border=None):
        cell_range = f'{start_col}{row}:{end_col}{row}'
        self.ws.merge_cells(cell_range)
        cell = self.ws[f'{start_col}{row}']
        cell.value = value
        cell.font = font
        if fill:
            cell.fill = fill
        if alignment:
            cell.alignment = alignment
        if border:
            for col in range(ord(start_col), ord(end_col) + 1):
                self.ws[f'{chr(col)}{row}'].border = border

    def create_main_title(self, budget, duration, delivery):
        self._merge_and_style(
            row=1, start_col='A', end_col='G',
            value=f'{self.project_name}',
            font=Font(name='微软雅黑', bold=True, size=18, color=COLORS['main_title_font']),
            fill=PatternFill(start_color=COLORS['main_title_bg'],
                            end_color=COLORS['main_title_bg'], fill_type='solid'),
            alignment=Alignment(horizontal='center', vertical='center')
        )
        self.ws.row_dimensions[1].height = 38

        summary = f'项目预算：{budget}    |    开发周期：{duration}    |    交付时间：{delivery}'
        self._merge_and_style(
            row=2, start_col='A', end_col='G', value=summary,
            font=Font(name='微软雅黑', size=11),
            alignment=Alignment(horizontal='center', vertical='center')
        )
        self.ws.row_dimensions[2].height = 26
        self.current_row = 3

    def create_header(self):
        headers = ['功能模块', '功能名称', '功能说明', '功能价值与业务流程', '工期(天)', '费用(元)', '重要度']
        font = Font(name='微软雅黑', bold=True, size=11, color=COLORS['header_font'])
        fill = PatternFill(start_color=COLORS['header_bg'],
                          end_color=COLORS['header_bg'], fill_type='solid')
        align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx, header in enumerate(headers, 1):
            cell = self.ws.cell(row=3, column=col_idx, value=header)
            cell.font = font
            cell.fill = fill
            cell.alignment = align
            cell.border = THIN_BORDER

        self.ws.row_dimensions[3].height = 28
        self.current_row = 4

    def add_module(self, module_name, module_number, features):
        cn_nums = ['一', '二', '三', '四', '五', '六', '七', '八', '九', '十']
        prefix = cn_nums[module_number - 1] if module_number <= 10 else str(module_number)
        module_title = f'{prefix}、{module_name}'

        module_start_row = self.current_row

        for idx, f in enumerate(features, 1):
            row_num = self.current_row
            self.ws.cell(row=row_num, column=1).border = THIN_BORDER

            feature_no = f'{module_number}.{idx}'
            cells = [
                (2, f'{feature_no} {f["name"]}', 'normal'),
                (3, f.get('description', ''), 'normal'),
                (4, f.get('value_flow', ''), 'normal'),
                (5, f.get('days', 0), 'center'),
                (6, f.get('cost', 0), 'center'),
                (7, f.get('priority', 'P1'), 'center'),
            ]
            for col, value, align_kind in cells:
                cell = self.ws.cell(row=row_num, column=col, value=value)
                if align_kind == 'center':
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical='center', wrap_text=True)
                cell.border = THIN_BORDER

            self.ws.row_dimensions[row_num].height = 60
            self.current_row += 1

        feature_end_row = self.current_row - 1

        # 模块名合并到 A 列
        if feature_end_row >= module_start_row:
            self.ws.merge_cells(f'A{module_start_row}:A{feature_end_row}')
            cell = self.ws.cell(row=module_start_row, column=1, value=module_title)
            cell.font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['module_title_font'])
            cell.fill = PatternFill(start_color=COLORS['module_title_bg'],
                                    end_color=COLORS['module_title_bg'], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = THIN_BORDER

        # 模块小计行
        sub_fill = PatternFill(start_color=COLORS['subtotal_bg'],
                              end_color=COLORS['subtotal_bg'], fill_type='solid')
        sub_font = Font(name='微软雅黑', bold=True, color=COLORS['subtotal_font'])

        sub_a = self.ws.cell(row=self.current_row, column=1, value='小计')
        sub_a.font = sub_font; sub_a.fill = sub_fill; sub_a.border = THIN_BORDER
        sub_a.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(2, 5):
            c = self.ws.cell(row=self.current_row, column=col)
            c.fill = sub_fill; c.border = THIN_BORDER

        sub_e = self.ws.cell(row=self.current_row, column=5,
                            value=f'=SUM(E{module_start_row}:E{feature_end_row})')
        sub_e.font = sub_font; sub_e.fill = sub_fill; sub_e.border = THIN_BORDER
        sub_e.alignment = Alignment(horizontal='center', vertical='center')

        sub_f = self.ws.cell(row=self.current_row, column=6,
                            value=f'=SUM(F{module_start_row}:F{feature_end_row})')
        sub_f.font = sub_font; sub_f.fill = sub_fill; sub_f.border = THIN_BORDER
        sub_f.alignment = Alignment(horizontal='center', vertical='center')
        sub_f.number_format = '#,##0'

        sub_g = self.ws.cell(row=self.current_row, column=7)
        sub_g.fill = sub_fill; sub_g.border = THIN_BORDER

        self.subtotal_rows.append(self.current_row)
        self.ws.row_dimensions[self.current_row].height = 26
        self.current_row += 1

    def create_total_row(self, summary_text=None):
        total_fill = PatternFill(start_color=COLORS['total_bg'],
                                end_color=COLORS['total_bg'], fill_type='solid')
        total_font = Font(name='微软雅黑', bold=True, size=12, color=COLORS['total_font'])

        a = self.ws.cell(row=self.current_row, column=1, value='项目总计')
        a.font = total_font; a.fill = total_fill; a.border = THICK_BORDER
        a.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(2, 4):
            c = self.ws.cell(row=self.current_row, column=col)
            c.fill = total_fill; c.border = THICK_BORDER

        d = self.ws.cell(row=self.current_row, column=4, value=summary_text or '')
        d.font = Font(name='微软雅黑', bold=True, color=COLORS['total_font'])
        d.fill = total_fill; d.border = THICK_BORDER
        d.alignment = Alignment(vertical='center', wrap_text=True)

        if self.subtotal_rows:
            refs = ','.join([f'E{r}' for r in self.subtotal_rows])
            e = self.ws.cell(row=self.current_row, column=5, value=f'=SUM({refs})')
        else:
            e = self.ws.cell(row=self.current_row, column=5, value=0)
        e.font = total_font; e.fill = total_fill; e.border = THICK_BORDER
        e.alignment = Alignment(horizontal='center', vertical='center')

        if self.subtotal_rows:
            refs = ','.join([f'F{r}' for r in self.subtotal_rows])
            f = self.ws.cell(row=self.current_row, column=6, value=f'=SUM({refs})')
        else:
            f = self.ws.cell(row=self.current_row, column=6, value=0)
        f.font = total_font; f.fill = total_fill; f.border = THICK_BORDER
        f.alignment = Alignment(horizontal='center', vertical='center')
        f.number_format = '#,##0'

        g = self.ws.cell(row=self.current_row, column=7)
        g.fill = total_fill; g.border = THICK_BORDER

        self.ws.row_dimensions[self.current_row].height = 32
        self.current_row += 1

    def add_section(self, title, content):
        self.current_row += 1
        self._merge_and_style(
            row=self.current_row, start_col='A', end_col='G', value=title,
            font=Font(name='微软雅黑', bold=True, size=12, color=COLORS['section_title_font']),
            fill=PatternFill(start_color=COLORS['section_title_bg'],
                            end_color=COLORS['section_title_bg'], fill_type='solid'),
            alignment=Alignment(vertical='center', indent=1)
        )
        self.ws.row_dimensions[self.current_row].height = 26
        self.current_row += 1

        for line in content:
            self._merge_and_style(
                row=self.current_row, start_col='A', end_col='G', value=line,
                font=Font(name='微软雅黑', size=10),
                alignment=Alignment(vertical='center', wrap_text=True, indent=1)
            )
            self.ws.row_dimensions[self.current_row].height = 22
            self.current_row += 1

    def save(self, filename):
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        self.wb.save(filename)
        print(f'[OK] 报价单已保存：{filename}')
        return filename


# ==================== 项目专属配置 ====================

def create_aidui_quote():
    """爱对 - 多平台店铺财务自动对账系统 报价单"""

    g = ProjectQuoteGenerator('爱对 — 多平台店铺财务自动对账系统  报价单',
                              sheet_name='报价明细')

    g.create_main_title(
        budget='5 万元（含税）',
        duration='20 个工作日（4 周）',
        delivery='签约后 28 自然日内'
    )
    g.create_header()

    # ============ 模块1：商品管理（2.5 天 / 7000 元） ============
    g.add_module(
        module_name='商品管理\n(需求 #1 #2 #3 #4)',
        module_number=1,
        features=[
            {
                'name': '商品资料维护',
                'description': '维护款式编码 / 商品编码 / 商品名称 / 品类等基础字段，支持 Excel 模板下载、批量导入、在线编辑、关键词搜索',
                'value_flow': '价值：建立商品主数据，与对账/利润分析联动\n流程：下载模板 → 填写 → 上传 → 自动去重合并',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '商品成本管理',
                'description': '按期间维护「商品成本 + 标费 + 辅料 = 总成本」，支持批量导入、在线编辑、按 SKU 检索',
                'value_flow': '价值：成本数据从 Excel 散落集中到系统，与对账自动联动\n流程：录入/导入 → 对账时按期间+SKU匹配 → 自动算真实利润',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '成本版本控制',
                'description': '成本调整后不影响已生成的历史报表，新对账才用新成本（关键需求）',
                'value_flow': '价值：保障历史数据稳定，避免追溯混乱\n流程：变更前的报表生成快照 → 变更只影响后续',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '成本修改记录',
                'description': '记录每次成本变更的操作人/时间/前后值，支持按操作类型筛选、Excel 导出',
                'value_flow': '价值：便于审计追溯与责任划分\n流程：编辑保存 → 自动写日志 → 审计查询',
                'days': 0.5, 'cost': 1500, 'priority': 'P1'
            },
        ]
    )

    # ============ 模块2：订单与对账引擎（4.5 天 / 12000 元） ============
    g.add_module(
        module_name='订单管理\n与对账引擎\n(需求 #5 #7 #8 #9)',
        module_number=2,
        features=[
            {
                'name': '7 个国内平台账单解析',
                'description': '抖音、淘宝/天猫、拼多多、快手、视频号小店、微信小店、小红书 7 个平台原始账单 Excel 解析，统一映射到标准列名',
                'value_flow': '价值：跨平台一套引擎，新平台只加映射文件\n流程：上传 .xlsx → 适配器映射 → 标准化输出',
                'days': 1.5, 'cost': 4000, 'priority': 'P0'
            },
            {
                'name': '聚水潭数据导入',
                'description': '聚水潭"销售主题分析-明细(订单商品)"导出文件解析，支持多月数据、模糊 sheet 名匹配',
                'value_flow': '价值：与平台账单交叉对账的基础数据\n流程：上传 → 按订单号聚合 → 输出 JstOrder',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '5 桶差异分析引擎',
                'description': '订单按 matched / duplicated / missing_in_jst / missing_in_platform / profit_anomaly 5 类自动分桶，支持 AI 提示一句话定位差异原因',
                'value_flow': '价值：人工 3 天对账 → 系统 30 秒\n流程：对账 → KPI 卡 + 差异表 + 一句话提示',
                'days': 2, 'cost': 5500, 'priority': 'P0'
            },
        ]
    )

    # ============ 模块3：海外平台与汇率管理（2.5 天 / 6500 元） ============
    g.add_module(
        module_name='海外平台\n与汇率管理\n(需求 #6 #10 #11 #12)',
        module_number=3,
        features=[
            {
                'name': '4 个海外平台适配器',
                'description': '俄罗斯 OZON / 俄罗斯 Wildberries / 英国 TikTok Shop / 美国 Amazon 4 个平台账单字段映射，币种自动识别',
                'value_flow': '价值：解决海外多平台对账空白\n流程：选海外平台 → 上传账单 → 自动按币种处理',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '汇率引擎与维护页',
                'description': '按客户规则「账单期间 → 取次月1日汇率」自动换算 USD/RUB/GBP 到人民币，支持按月维护汇率，节假日异常说明',
                'value_flow': '价值：海外多币种统一为 CNY 核算\n流程：维护汇率 → 对账自动换算 → 报表统一展示',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '币种与店铺联动',
                'description': '店铺档案带默认币种，平台切换时自动联动，对账结果带原币/CNY 双口径展示',
                'value_flow': '价值：跨币种透明可追溯\n流程：店铺配置 → 对账 → 双币种展示',
                'days': 0.5, 'cost': 1500, 'priority': 'P1'
            },
        ]
    )

    # ============ 模块4：利润核算中心（3.5 天 / 9500 元） ============
    g.add_module(
        module_name='利润核算中心\n(数据归集/分配/分析)',
        module_number=4,
        features=[
            {
                'name': '数据归集',
                'description': '按费用类型（推广费/运费险/平台服务费/红包/补贴等 10 类）归集到组织/店铺/平台单号维度，支持 Excel 批量导入',
                'value_flow': '价值：公共费用集中管理\n流程：录入/导入 → 按类型分类 → 等待分配',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '分配标准',
                'description': '5 种分配方式（按收入/按件数/按订单数/平均/直挂订单）×3 种范围（组织/店铺/订单商品）自由组合，支持优先级匹配',
                'value_flow': '价值：分摊规则透明可配置\n流程：定义规则 → 分配引擎按优先级匹配 → 自动分摊',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '三层分配引擎',
                'description': '组织 → 店铺 → 订单商品 三层分配，含 4 个聚合视图（按组织/店铺/商品/订单），未分配原因可视化',
                'value_flow': '价值：每分钱费用都能落到 SKU\n流程：自动按规则匹配 → 分摊 → 多维上卷',
                'days': 1.5, 'cost': 4000, 'priority': 'P0'
            },
            {
                'name': '利润分析表（多维）',
                'description': '商品 / 店铺 / 平台 / 品类 4 维度切换，支持 KPI 卡片 + 上卷/下钻',
                'value_flow': '价值：一键看清各维度盈利能力\n流程：选维度 → 自动上卷 → 排序/筛选',
                'days': 0.5, 'cost': 1500, 'priority': 'P1'
            },
        ]
    )

    # ============ 模块5：数据中心 7 张报表（3.5 天 / 9000 元） ============
    g.add_module(
        module_name='数据中心\n(7 张报表 - 需求 #13~#16)',
        module_number=5,
        features=[
            {
                'name': '账单汇总/明细表',
                'description': '严格按客户《系统初稿模板5.22》表头：平台/店铺/月初余额/收入项(4)/支出项(6)/月末余额。明细级到订单/款式编码',
                'value_flow': '价值：与客户既有 Excel 报表表头逐字一致\n流程：对账后自动生成，可一键导出 Excel',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '店铺利润表',
                'description': '严格按客户表头：销售收入(4) + 销售成本(7) + 销售费用(7) + 店铺利润 + 毛利率 + 退货率 + 确收率',
                'value_flow': '价值：店铺老板视角的核心利润看板\n流程：对账+分配后自动生成',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '商品利润表',
                'description': '15 列（订单号/款式/商品编码/商品名称/品类/数量/单价/销售金额/成本/标费/辅料/毛利润/毛利率/备注）+ 右侧智能分析区（关键指标卡 + 12 月柱状图）',
                'value_flow': '价值：SKU 级利润排行 + 月度趋势\n流程：对账+成本+分配 → 自动生成',
                'days': 1, 'cost': 2500, 'priority': 'P0'
            },
            {
                'name': '应收汇总/明细表',
                'description': '店铺/商品维度：期初结余 → 本期应收 → 本期核销 → 期末结余（数量+金额）',
                'value_flow': '价值：跨月应收余额追踪\n流程：对账后自动生成',
                'days': 0.5, 'cost': 1000, 'priority': 'P1'
            },
            {
                'name': '差异分析表（对账主入口）',
                'description': '上传账单 + 对账 + KPI 卡片 + 5 桶筛选 + 行下钻原始流水 + 月度公共扣费面板',
                'value_flow': '价值：对账核心工作台\n流程：上传 → 对账 → 看 KPI → 筛差异 → 下钻',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
        ]
    )

    # ============ 模块6：店铺管理与系统设置（1.5 天 / 3500 元） ============
    g.add_module(
        module_name='店铺管理\n与系统设置\n(需求 #17~#20)',
        module_number=6,
        features=[
            {
                'name': '店铺/平台/币种配置',
                'description': '维护店铺名称/所属平台/默认币种/营业状态/结算规则；支持按平台筛选、一键停用',
                'value_flow': '价值：多店铺集中管理\n流程：录入 → 关联平台 → 对账时自动绑定',
                'days': 0.5, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '角色/权限/数据备份',
                'description': '5 个预置角色（超管/财务主管/财务/运营/只读）+ 权限矩阵展示 + 一键导出全部数据 JSON 备份',
                'value_flow': '价值：满足审计与多人协作\n流程：分配角色 → 操作受控 → 定期备份',
                'days': 0.5, 'cost': 1000, 'priority': 'P1'
            },
            {
                'name': 'Excel 导出与模板下载',
                'description': '所有报表支持导出 Excel（多 sheet + 列宽 + 货币格式），关键页提供导入模板下载',
                'value_flow': '价值：与现有财务流程无缝衔接\n流程：报表页点导出 → 自动生成 .xlsx',
                'days': 0.5, 'cost': 1000, 'priority': 'P0'
            },
        ]
    )

    # ============ 模块7：测试/部署/培训（2 天 / 2500 元） ============
    g.add_module(
        module_name='测试 / 部署\n/ 培训',
        module_number=7,
        features=[
            {
                'name': '回归测试与质量验收',
                'description': '使用真实抖音账单做回归测试，覆盖核心对账与利润核算流程',
                'value_flow': '价值：保障对账准确性 ≥ 95%\n流程：测试 → 修正 → 复测',
                'days': 1, 'cost': 1500, 'priority': 'P0'
            },
            {
                'name': '上线部署',
                'description': '部署到客户指定环境（公网域名或内部服务器均可），配置域名访问',
                'value_flow': '价值：可访问的线上系统\n流程：构建 → 上传 → 域名解析 → 验证',
                'days': 0.5, 'cost': 500, 'priority': 'P0'
            },
            {
                'name': '现场培训与文档',
                'description': '1 次现场/远程培训 + 用户操作手册 + 开发者扩展指南 + 培训视频',
                'value_flow': '价值：5 人内财务团队 1 小时上手\n流程：演示 → 答疑 → 留资料',
                'days': 0.5, 'cost': 500, 'priority': 'P0'
            },
        ]
    )

    # 项目总计
    g.create_total_row(
        summary_text='共 7 大模块，22 项核心功能；满足客户需求表全部 20 项 + 系统初稿模板全部 4 张报表'
    )

    # ============ 说明区 ============

    g.add_section('📋 项目说明', [
        '本项目为客户《财务自动对账系统-需求表》20 项需求 + 《系统初稿模板5.22》4 张报表的完整实现。',
        '系统采用浏览器端纯 Web 方案，财务无需安装任何软件，打开网页即可使用。',
        '11 个平台（7 国内 + 4 海外）统一管理，新增平台扩展灵活。',
    ])

    g.add_section('💰 费用说明', [
        '• 报价含税（增值税专票），合计 50,000 元整',
        '• 含：需求确认 / 全部功能开发 / 测试 / 部署上线 / 1 次培训 / 3 个月免费维护',
        '• 不含：第三方平台费用（如域名/服务器）、二次开发后端 API、对接金蝶/用友 ERP',
        '• 付款方式建议：合同签订付 30%、对账主流程交付付 40%、验收完成付 30%',
    ])

    g.add_section('⏱️ 开发周期（20 工作日 / 4 周）', [
        '阶段一 [Day 1-4 / 第 1 周]：需求确认 + 商品管理 3 页（资料/成本/历史）',
        '阶段二 [Day 5-10 / 第 2 周]：对账引擎 + 7 国内平台 + 聚水潭 + 差异分析表',
        '阶段三 [Day 11-16 / 第 3 周]：汇率/分配引擎 + 利润核算 + 7 张报表',
        '阶段四 [Day 17-20 / 第 4 周]：4 海外平台 + 系统设置 + 测试部署 + 培训交付',
    ])

    g.add_section('🎯 功能优先级说明', [
        'P0（必做）：商品资料/成本/审计、对账引擎、5 桶差异、汇率引擎、4 张核心报表、店铺配置、Excel 导出',
        'P1（重要）：4 维利润分析、应收报表、角色权限、智能分析区',
    ])

    g.add_section('🔧 系统说明', [
        '• 浏览器端纯 Web 系统，无需安装任何客户端',
        '• 支持 Chrome / Edge 等主流浏览器',
        '• 数据保存在客户自有环境，不上传第三方服务',
        '• 支持上线到客户指定地址（公网域名 / 内部服务器均可）',
        '• 满足客户《系统初稿模板5.22》全部表头规范',
    ])

    g.add_section('📦 交付清单', [
        '• 完整系统（含历史变更记录可追溯）',
        '• 上线就绪的可执行版本 + 上线说明文档',
        '• 用户操作手册（PDF/在线）',
        '• 系统扩展指南（新增平台/报表/字段）',
        '• 演示数据（11 平台 + 全模块种子数据）',
        '• 1 次现场/远程培训 + 录屏回放',
    ])

    g.add_section('🛡️ 售后服务', [
        '• 交付后 3 个月免费维护（含 bug 修复、浏览器兼容、文档勘误）',
        '• 1 次远程答疑',
        '• 季度免费版本升级（已知问题修复、安全补丁）',
        '• 重大版本升级 8 折优惠',
        '• 新增国内平台：1500 元 / 个；新增海外平台：2500 元 / 个',
        '• 新增自定义报表：2000-3500 元 / 张',
    ])

    g.add_section('⚠️ 风险与待决策事项', [
        '• 客户需要在第 1 周提供 3-5 个真实账单样本以提前调试适配',
        '• 海外平台账单字段需客户提供 1 份样本，否则用通用模板',
        '• 数据存储默认浏览器本地，如需多端同步需追加服务端能力（单独评估）',
        '• 角色/权限/备份目前为前端展示，生产级权限控制需服务端配套',
    ])

    out_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # 方案与报价 目录
    out_path = os.path.join(out_dir, '爱对-报价单.xlsx')
    g.save(out_path)
    return out_path


if __name__ == '__main__':
    create_aidui_quote()
